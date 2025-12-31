import openpyxl
from openpyxl.styles import Font, Alignment
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.workbook.workbook import Workbook
from typing import Dict, Any, List, Optional, Tuple


class ApplicantExcelWriter:
    """Excel writer for applicant data management."""

    # Constants
    WORK_EXPERIENCE_COUNT = 5
    WORK_EXPERIENCE_BASE_COLUMNS = ['J', 'P', 'V', 'AB', 'AH']
    WORK_EXPERIENCE_FIELDS = ['start_date', 'end_date', 'company_name',
                               'final_department', 'final_position', 'salary']
    WORK_EXPERIENCE_LABELS = ['경력 1', '경력 2', '경력 3', '경력 4', '경력 5']
    WORK_EXPERIENCE_FIELD_NAMES = ['입사년월', '퇴사년월', '회사명',
                                    '최종부서명', '최종직위', '연봉(천원)']

    MAX_COLUMN_WIDTH = 30
    COLUMN_PADDING = 2
    SHEET_TITLE = "지원자 데이터"

    def __init__(self, excel_path: str):
        """
        Initialize ApplicantExcelWriter.

        Args:
            excel_path: Path to the Excel file
        """
        self.excel_path = excel_path
        self.column_mapping = self._create_column_mapping()
        
    def _create_column_mapping(self) -> Dict[str, str]:
        """
        Define the mapping between JSON paths and Excel columns.

        Returns:
            Dictionary mapping JSON paths to Excel column letters
        """
        mapping = {
            'applicant_number': 'A',
            'applicant_name': 'B',
            'application_date': 'C',
            'affiliation': 'D',
            'application_field': 'E',
            'basic_info.birth_year': 'F',
            'basic_info.gender': 'G',
            'basic_info.final_education_school': 'H',
            'basic_info.final_education_degree': 'I',
        }

        mapping.update(self._create_work_experience_mapping())
        return mapping

    def _create_work_experience_mapping(self) -> Dict[str, str]:
        """
        Create column mapping for work experience fields.

        Returns:
            Dictionary mapping work experience JSON paths to Excel columns
        """
        mapping = {}
        for exp_idx in range(self.WORK_EXPERIENCE_COUNT):
            base_col_idx = self._column_to_index(
                self.WORK_EXPERIENCE_BASE_COLUMNS[exp_idx]
            )
            for field_offset, field in enumerate(self.WORK_EXPERIENCE_FIELDS):
                col = self._index_to_column(base_col_idx + field_offset)
                mapping[f'work_experience[{exp_idx}].{field}'] = col

        return mapping
    
    @staticmethod
    def _column_to_index(col: str) -> int:
        """
        Convert Excel column letter to 0-based index.

        Args:
            col: Excel column letter (e.g., 'A', 'B', 'AA')

        Returns:
            0-based column index
        """
        result = 0
        for char in col:
            result = result * 26 + (ord(char) - ord('A') + 1)
        return result - 1

    @staticmethod
    def _index_to_column(idx: int) -> str:
        """
        Convert 0-based index to Excel column letter.

        Args:
            idx: 0-based column index

        Returns:
            Excel column letter (e.g., 'A', 'B', 'AA')
        """
        col = ''
        idx += 1
        while idx > 0:
            idx -= 1
            col = chr(idx % 26 + ord('A')) + col
            idx //= 26
        return col
    
    @staticmethod
    def _get_nested_value(data: Dict[str, Any], path: str) -> Optional[Any]:
        """
        Extract value from nested dictionary using dot notation and array indices.

        Args:
            data: Source dictionary
            path: Path to the value (e.g., 'basic_info.name' or 'work_experience[0].company')

        Returns:
            Value at the specified path, or None if not found
        """
        keys = path.replace('[', '.').replace(']', '').split('.')
        value = data

        for key in keys:
            if key.isdigit():
                idx = int(key)
                if isinstance(value, list) and idx < len(value):
                    value = value[idx]
                else:
                    return None
            else:
                value = value.get(key) if isinstance(value, dict) else None

            if value is None:
                return None

        return value
    
    def _get_basic_headers(self) -> Dict[str, str]:
        """
        Get basic information headers.

        Returns:
            Dictionary mapping cell references to header text
        """
        return {
            'A1': '지원자 번호',
            'B1': '지원자명',
            'C1': '지원일',
            'D1': '소속',
            'E1': '지원분야/공고',
            'F1': '출생년도',
            'G1': '성별',
            'H1': '최종학력(학교)',
            'I1': '최종학력(학사 등)',
        }

    def _get_work_experience_headers(self) -> Dict[str, str]:
        """
        Get work experience headers.

        Returns:
            Dictionary mapping cell references to header text
        """
        headers = {}
        for exp_idx in range(self.WORK_EXPERIENCE_COUNT):
            label = self.WORK_EXPERIENCE_LABELS[exp_idx]
            base_col = self.WORK_EXPERIENCE_BASE_COLUMNS[exp_idx]
            base_idx = self._column_to_index(base_col)

            for field_offset, field_name in enumerate(self.WORK_EXPERIENCE_FIELD_NAMES):
                col = self._index_to_column(base_idx + field_offset)
                headers[f'{col}1'] = f'{label} {field_name}'

        return headers

    def _apply_header_style(self, ws: Worksheet, headers: Dict[str, str]) -> None:
        """
        Apply styling to header cells.

        Args:
            ws: Worksheet to apply styling to
            headers: Dictionary mapping cell references to header text
        """
        header_font = Font(bold=True)
        header_alignment = Alignment(horizontal='center')

        for cell_ref, header_text in headers.items():
            cell = ws[cell_ref]
            cell.value = header_text
            cell.font = header_font
            cell.alignment = header_alignment

    def _adjust_column_widths(self, ws: Worksheet) -> None:
        """
        Adjust column widths based on content.

        Args:
            ws: Worksheet to adjust
        """
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter

            for cell in column:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))

            adjusted_width = min(
                max_length + self.COLUMN_PADDING,
                self.MAX_COLUMN_WIDTH
            )
            ws.column_dimensions[column_letter].width = adjusted_width

    def create_template(self) -> None:
        """Create a new Excel file with headers."""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = self.SHEET_TITLE

        headers = {**self._get_basic_headers(), **self._get_work_experience_headers()}
        self._apply_header_style(ws, headers)
        self._adjust_column_widths(ws)

        wb.save(self.excel_path)
        print(f"Template created: {self.excel_path}")
    
    def _load_or_create_workbook(self) -> Tuple[Workbook, Worksheet]:
        """
        Load existing workbook or create new template.

        Returns:
            Tuple of (Workbook, Worksheet)
        """
        try:
            wb = openpyxl.load_workbook(self.excel_path)
            ws = wb.active
        except FileNotFoundError:
            print("Excel file not found. Creating new template...")
            self.create_template()
            wb = openpyxl.load_workbook(self.excel_path)
            ws = wb.active

        return wb, ws

    def _generate_applicant_number(self, row_number: int) -> str:
        """
        Generate sequential applicant number.

        Args:
            row_number: Current row number in Excel

        Returns:
            Generated applicant number
        """
        return str(row_number - 1)

    def _write_applicant_row(
        self,
        ws: Worksheet,
        row_number: int,
        json_data: Dict[str, Any]
    ) -> None:
        """
        Write applicant data to a specific row.

        Args:
            ws: Target worksheet
            row_number: Target row number
            json_data: Applicant data
        """
        cell_alignment = Alignment(horizontal='center', vertical='center')

        for json_path, excel_col in self.column_mapping.items():
            value = self._get_nested_value(json_data, json_path)
            cell_ref = f'{excel_col}{row_number}'

            cell = ws[cell_ref]
            cell.value = value if value is not None else ''
            cell.alignment = cell_alignment

    def append_applicant(self, json_data: Dict[str, Any]) -> int:
        """
        Append a new applicant to the Excel file.

        Args:
            json_data: Applicant data in dictionary format

        Returns:
            Row number where the applicant was added
        """
        wb, ws = self._load_or_create_workbook()
        next_row = ws.max_row + 1

        if not json_data.get('applicant_number'):
            json_data['applicant_number'] = self._generate_applicant_number(next_row)

        self._write_applicant_row(ws, next_row, json_data)

        wb.save(self.excel_path)
        print(f"Applicant added at row {next_row}: {json_data.get('applicant_name', 'Unknown')}")
        return next_row
    
    def batch_append(self, json_data_list: List[Dict[str, Any]]) -> None:
        """
        Append multiple applicants at once (more efficient than individual appends).

        Args:
            json_data_list: List of applicant data dictionaries
        """
        if not json_data_list:
            print("No applicants to add.")
            return

        wb, ws = self._load_or_create_workbook()
        start_row = ws.max_row + 1

        for idx, json_data in enumerate(json_data_list):
            current_row = start_row + idx

            if not json_data.get('applicant_number'):
                json_data['applicant_number'] = self._generate_applicant_number(current_row)

            self._write_applicant_row(ws, current_row, json_data)

        wb.save(self.excel_path)
        end_row = start_row + len(json_data_list) - 1
        print(f"Batch added {len(json_data_list)} applicants (rows {start_row}-{end_row})")
